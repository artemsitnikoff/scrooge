import json
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from models import WeighingRecord

# Маппинг заголовков Excel → поля модели
_HEADER_MAP = {
    "id": "id",
    "дата въезда": "dateBefore",
    "дата выезда": "dateAfter",
    "datebefore": "dateBefore",
    "dateafter": "dateAfter",
    "госномер": "registrationNumber",
    "номер тс": "registrationNumber",
    "registrationnumber": "registrationNumber",
    "тип тс": "garbageTruckType",
    "garbagetrucktype": "garbageTruckType",
    "марка тс": "garbageTruckBrand",
    "garbagetruckbrand": "garbageTruckBrand",
    "модель тс": "garbageTruckModel",
    "garbagetruckmodel": "garbageTruckModel",
    "контрагент": "companyName",
    "companyname": "companyName",
    "инн": "companyInn",
    "companyinn": "companyInn",
    "кпп": "companyKpp",
    "companykpp": "companyKpp",
    "вес въезд": "weightBefore",
    "вес на въезде": "weightBefore",
    "weightbefore": "weightBefore",
    "вес выезд": "weightAfter",
    "вес на выезде": "weightAfter",
    "weightafter": "weightAfter",
    "вес водителя": "weightDriver",
    "weightdriver": "weightDriver",
    "коэффициент": "coefficient",
    "coefficient": "coefficient",
    "вес мусора": "garbageWeight",
    "garbageweight": "garbageWeight",
    "тип отходов": "garbageType",
    "garbagetype": "garbageType",
    "код фкко": "codeFKKO",
    "codefkko": "codeFKKO",
    "наименование фкко": "nameFKKO",
    "namefkko": "nameFKKO",
}


def parse_file(file_path: str) -> tuple[list[dict], list[str]]:
    ext = Path(file_path).suffix.lower()
    if ext == ".json":
        return _parse_json(file_path)
    elif ext in (".xlsx", ".xls"):
        return _parse_xlsx(file_path)
    else:
        return [], [f"Неподдерживаемый формат файла: {ext}. Ожидается .xlsx или .json"]


def _parse_json(file_path: str) -> tuple[list[dict], list[str]]:
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Поддерживаем как массив, так и объект с полем weightControls
    if isinstance(data, dict):
        raw_records = data.get("weightControls", data.get("records", []))
    elif isinstance(data, list):
        raw_records = data
    else:
        return [], ["JSON должен содержать массив записей или объект с полем weightControls"]

    return _validate_records(raw_records)


def _parse_xlsx(file_path: str) -> tuple[list[dict], list[str]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], ["Файл пуст"]

    # Определяем маппинг колонок по заголовкам первой строки
    headers = rows[0]
    col_map: dict[int, str] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in _HEADER_MAP:
            col_map[i] = _HEADER_MAP[key]

    if not col_map:
        return [], ["Не удалось определить колонки. Проверьте заголовки в файле."]

    raw_records = []
    for row in rows[1:]:
        record = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                record[field_name] = str(val).strip()
        if record:
            raw_records.append(record)

    return _validate_records(raw_records)


def _validate_records(raw_records: list[dict]) -> tuple[list[dict], list[str]]:
    valid = []
    errors = []

    for i, raw in enumerate(raw_records, start=1):
        try:
            record = WeighingRecord(**raw)
            valid.append(record.model_dump(exclude_none=True))
        except ValidationError as e:
            for err in e.errors():
                field = ".".join(str(x) for x in err["loc"])
                errors.append(f"Строка {i}, поле «{field}»: {err['msg']}")

    return valid, errors
